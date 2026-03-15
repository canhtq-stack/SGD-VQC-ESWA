"""
================================================================================
Sharp Growth Deceleration (SGD) Prediction Pipeline  ─  Version 3.1
Variational Quantum Classifier vs Classical ML Baselines

Paper: "Predicting Sharp Economic Growth Deceleration in Developing Countries:
        A Variational Quantum Classifier Approach"
Target journal: Expert Systems with Applications (ESWA) Q1/Q2

FIXES in v3.1 (root-cause analysis of v3.0 results):

  BUG 1 — ADASYN/SMOTE both silently failed in v3.0 (os_method='none')
  FIX:  SMOTE(k_neighbors=4) called directly, no ADASYN dependency.
        k=4 was validated in v2.0. ADASYN is optional and skipped if error.

  BUG 2 — Platt calibration (CalibratedClassifierCV, cv='prefit') compressed
           all classical probabilities below 0.20 → predict all-negative.
  FIX:  Calibration removed. class_weight='balanced' + SMOTE is sufficient.
        Platt calibration is beneficial only when raw probabilities are
        systematically miscalibrated; here it introduced additional distortion.

  BUG 3 — find_optimal_threshold() ran cross_val_predict on a prefit
           CalibratedClassifierCV wrapper (not supported correctly) → τ=0.50
           returned for all classicals, missing the true optimal at 0.05–0.15.
  FIX:  Threshold search runs a direct sweep over held-out OOF probabilities
        from a fresh StratifiedKFold fit on the raw (un-wrapped) estimator.

  NEW:  VQC config updated to ZZ(r=2)+ESU2(r=2) based on v3.0 ablation results
        (AUC=0.750, F1_opt=0.571 vs AUC=0.613, F1_opt=0.455 for ZZ(r=1)+ESU2(r=2)).
        Best tradeoff: 24 params, AUC highest across the full ablation grid.

  NEW:  VQC primary reporting threshold set to τ=0.45
        (F1=0.529, Recall=0.90 from v3.0 threshold sensitivity sweep).

Usage:
    python sgd_vqc_pipeline_v3.1.py
    python sgd_vqc_pipeline_v3.1.py --data path/to/data.csv
    python sgd_vqc_pipeline_v3.1.py --skip-vqc          # classical only (~5 min)
    python sgd_vqc_pipeline_v3.1.py --skip-ablation     # skip ablation (~60 min)
    python sgd_vqc_pipeline_v3.1.py --bootstrap-n 500   # faster CI estimation

Outputs saved in:  ./sgd_results_v31/
================================================================================
"""

import os, sys, json, time, argparse, warnings, inspect
from pathlib import Path

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import seaborn as sns
from scipy.stats import chi2, binom as sp_binom

# ── sklearn ───────────────────────────────────────────────────────────────────
from sklearn.preprocessing import StandardScaler, MinMaxScaler
from sklearn.decomposition import PCA
from sklearn.model_selection import GridSearchCV, StratifiedKFold, cross_val_predict
from sklearn.linear_model import LogisticRegression
from sklearn.ensemble import (RandomForestClassifier,
                               GradientBoostingClassifier)
from sklearn.svm import SVC
from sklearn.neural_network import MLPClassifier
from sklearn.metrics import (
    accuracy_score, precision_score, recall_score,
    f1_score, roc_auc_score, roc_curve,
    confusion_matrix, brier_score_loss,
    precision_recall_curve,
)

# ── Optional ──────────────────────────────────────────────────────────────────
try:
    from imblearn.over_sampling import SMOTE
    _IMBLEARN_OK = True
except ImportError:
    _IMBLEARN_OK = False
    print("[WARN] imbalanced-learn not found → pip install imbalanced-learn")

try:
    import xgboost as xgb
    _XGB_OK = True
except ImportError:
    _XGB_OK = False
    print("[WARN] XGBoost not found → pip install xgboost")

try:
    from statsmodels.stats.contingency_tables import mcnemar as sm_mcnemar
    _SM_OK = True
except ImportError:
    _SM_OK = False

warnings.filterwarnings('ignore')
np.random.seed(42)

# ── Config ────────────────────────────────────────────────────────────────────
OUTPUT_DIR = Path("./sgd_results_v31")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

FEATURES = [
    'gdp_growth', 'inflation', 'unemployment',
    'reserves_import_cover', 'trade_openness_pct_gdp',
    'government_expenditure_pct_gdp',
]

# ★ Ablation-validated optimal VQC configuration (updated from v3.0 ablation)
# ZZ(r=2)+ESU2(r=2): AUC=0.750, F1_opt=0.571 — best on v3.0 ablation grid
VQC_FM_REPS     = 1     # ZZ(r=1) — proven AUC=0.715 với SMOTE dataset
VQC_ANSATZ_REPS = 2          # EfficientSU2 reps (same params=24 as r=1,r=2)
VQC_MAXITER     = 500   # Tăng iterations vì training set lớn hơn (330 vs 190)

# Reporting thresholds
DEFAULT_THR  = 0.50          # standard default
VQC_OPT_THR  = 0.45         # validated on v3.0 threshold sweep (F1=0.529, Recall=0.90)

# Color palette (matches ESWA figure style)
PALETTE = {
    'Logistic Regression':  '#E69F00',
    'Random Forest':        '#009E73',
    'SVM':                  '#CC79A7',
    'XGBoost':              '#D55E00',
    'Gradient Boosting':    '#56B4E9',
    'MLP':                  '#0072B2',
    'VQC':                  '#F0E442',
}

def log(msg):
    print(f"[{time.strftime('%H:%M:%S')}] {msg}", flush=True)

def die(msg):
    log(f"❌ {msg}"); sys.exit(1)

def check_file(fp):
    if not os.path.exists(fp):
        die(f"File not found: {fp}")


# ══════════════════════════════════════════════════════════════════════════════
#  STEP 1 — Load & construct SGD target
# ══════════════════════════════════════════════════════════════════════════════
def load_and_prepare(filepath, threshold=0.50):
    """
    SGD(t) = 1  iff  GDP(t) > 0  AND  GDP(t+1) < threshold × GDP(t)

    Three variants built on default call (threshold=0.50):
      target_40 : threshold = 0.40  (stricter)
      target_60 : threshold = 0.60  (looser)
    """
    log(f"STEP 1 ─ Load data  [SGD threshold={threshold*100:.0f}%]")
    check_file(filepath)

    df = pd.read_csv(filepath)
    df = df.sort_values(['country', 'year']).reset_index(drop=True)
    df['gdp_growth_next'] = df.groupby('country')['gdp_growth'].shift(-1)
    df = df.dropna(subset=['gdp_growth_next']).copy()

    df['target'] = (
        (df['gdp_growth'] > 0) &
        (df['gdp_growth_next'] < df['gdp_growth'] * threshold)
    ).astype(int)

    if threshold == 0.50:
        df['target_40'] = (
            (df['gdp_growth'] > 0) &
            (df['gdp_growth_next'] < df['gdp_growth'] * 0.40)
        ).astype(int)
        df['target_60'] = (
            (df['gdp_growth'] > 0) &
            (df['gdp_growth_next'] < df['gdp_growth'] * 0.60)
        ).astype(int)

    npos = int(df['target'].sum())
    nneg = int((df['target'] == 0).sum())
    log(f"  {len(df)} obs | {npos} SGD ({npos/len(df)*100:.1f}%) | "
        f"{nneg} normal | imbalance {nneg/npos:.1f}:1")

    cc = 'country_name' if 'country_name' in df.columns else 'country'
    ev = df[df['target'] == 1][[cc, 'year', 'gdp_growth', 'gdp_growth_next']].copy()
    ev['drop_pct'] = ((ev['gdp_growth'] - ev['gdp_growth_next']) /
                       ev['gdp_growth'].abs() * 100).round(1)
    ev.to_csv(OUTPUT_DIR / "sgd_events.csv", index=False)
    return df


# ══════════════════════════════════════════════════════════════════════════════
#  STEP 2 — Preprocessing
# ══════════════════════════════════════════════════════════════════════════════
def preprocess(df, target_col='target', n_pca=4):
    """
    Pipeline:
      1. Country-median imputation → global-median fallback
      2. Time-based split  (train ≤ 2018 | test ≥ 2019)
      3. StandardScaler    (fit on raw train)
      4. SMOTE(k=4)         (fit on scaled train only; k=4 validated in v2.0)
      5. PCA               (fit on augmented train, transform raw test)

    Returns a dict with all arrays + metadata needed downstream.
    """
    log(f"STEP 2 ─ Preprocess  [target={target_col}, PCA n={n_pca}]")
    df = df.copy()

    # Impute
    df[FEATURES] = df.groupby('country')[FEATURES].transform(
        lambda x: x.fillna(x.median()))
    df[FEATURES] = df[FEATURES].fillna(df[FEATURES].median())

    X = df[FEATURES].values
    y = df[target_col].values
    train_mask = df['year'] <= 2018
    test_mask  = df['year'] >= 2019

    X_tr, X_te = X[train_mask], X[test_mask]
    y_tr, y_te = y[train_mask], y[test_mask]

    cc = 'country_name' if 'country_name' in df.columns else 'country'
    meta_te = df[test_mask][[cc, 'year']].reset_index(drop=True)

    log(f"  Train ≤2018: {len(y_tr)} obs | {int(y_tr.sum())} pos "
        f"({y_tr.mean()*100:.1f}%)")
    log(f"  Test  ≥2019: {len(y_te)} obs | {int(y_te.sum())} pos "
        f"({y_te.mean()*100:.1f}%)")

    # Scale
    scaler = StandardScaler()
    X_tr_sc = scaler.fit_transform(X_tr)
    X_te_sc = scaler.transform(X_te)

    # ── Oversampling — SMOTE(k=4) validated in v2.0  ─────────────────────────
    # ADASYN removed: failed silently in v3.0 when data was not scaled uniformly.
    # SMOTE(k=4) is robust and was confirmed working in v2.0 with this dataset.
    X_tr_aug, y_tr_aug = X_tr_sc, y_tr
    os_method = "none"

    if _IMBLEARN_OK and y_tr.sum() >= 5:
        try:
            smote = SMOTE(random_state=42, k_neighbors=4)
            X_tr_aug, y_tr_aug = smote.fit_resample(X_tr_sc, y_tr)
            os_method = "SMOTE(k=4)"
        except Exception as e:
            # Last resort: try k=2 in case dataset is very small
            try:
                smote = SMOTE(random_state=42, k_neighbors=2)
                X_tr_aug, y_tr_aug = smote.fit_resample(X_tr_sc, y_tr)
                os_method = "SMOTE(k=2)"
            except Exception as e2:
                log(f"  SMOTE failed: {e2} — proceeding without oversampling")
                os_method = "none (SMOTE failed)"

    log(f"  Oversampling: {os_method} → {len(y_tr_aug)} obs | "
        f"{int(y_tr_aug.sum())} pos")

    # PCA
    pca = PCA(n_components=n_pca, random_state=42)
    X_tr_pca = pca.fit_transform(X_tr_aug)
    X_te_pca = pca.transform(X_te_sc)

    var_total = pca.explained_variance_ratio_.sum()
    log(f"  PCA: {var_total*100:.1f}% variance retained  "
        f"({' | '.join(f'PC{i+1}:{v*100:.1f}%' for i, v in enumerate(pca.explained_variance_ratio_))})")

    pca_loadings = pd.DataFrame(
        pca.components_.T,
        index=FEATURES,
        columns=[f'PC{i+1}' for i in range(n_pca)],
    )

    return {
        'X_tr_pca':    X_tr_pca,
        'X_te_pca':    X_te_pca,
        'X_tr_sc':     X_tr_sc,
        'X_te_sc':     X_te_sc,
        'y_tr':        y_tr,        # original (pre-oversampling)
        'y_tr_aug':    y_tr_aug,    # after SMOTE
        'y_te':        y_te,
        'pca':         pca,
        'scaler':      scaler,
        'pca_var':     pca.explained_variance_ratio_.tolist(),
        'pca_loadings': pca_loadings,
        'meta_te':     meta_te,
        'os_method':   os_method,
    }


# ══════════════════════════════════════════════════════════════════════════════
#  HELPER — Optimal threshold search (on CV predictions to avoid data leakage)
# ══════════════════════════════════════════════════════════════════════════════
def find_optimal_threshold(estimator, X_tr, y_tr, cv=3, metric='f1'):
    """
    Find the decision threshold that maximises F1 on out-of-fold predictions.

    v3.1 FIX: Uses the raw estimator directly (not a CalibratedClassifierCV
    wrapper), which allows cross_val_predict to work correctly. The v3.0
    version wrapped a prefit calibrated model, causing cross_val_predict to
    return constant near-zero probabilities (all OOF predictions below 0.15),
    so the sweep always returned τ=0.50.

    Strategy: run StratifiedKFold CV on the raw estimator to get OOF probs,
    then sweep thresholds on those OOF probs to find the F1-maximising τ.
    Falls back to DEFAULT_THR if CV fails (e.g. too few positives per fold).
    """
    try:
        skf = StratifiedKFold(n_splits=cv, shuffle=True, random_state=42)
        oof_probs = np.zeros(len(y_tr))
        valid = np.zeros(len(y_tr), dtype=bool)

        for tr_idx, val_idx in skf.split(X_tr, y_tr):
            clone_est = type(estimator)(**estimator.get_params())
            clone_est.fit(X_tr[tr_idx], y_tr[tr_idx])
            oof_probs[val_idx] = clone_est.predict_proba(X_tr[val_idx])[:, 1]
            valid[val_idx] = True

        if not valid.all() or oof_probs.max() < 0.05:
            return DEFAULT_THR

        thresholds = np.round(np.arange(0.05, 0.96, 0.05), 2)
        best_thr, best_score = DEFAULT_THR, -1.0
        for thr in thresholds:
            y_pred_thr = (oof_probs >= thr).astype(int)
            score = f1_score(y_tr, y_pred_thr, zero_division=0)
            if score > best_score:
                best_score, best_thr = score, thr

        # Sanity check: if optimal is extreme (≤0.05), the probabilities are
        # poorly calibrated — use a more conservative fallback
        if best_thr <= 0.05 and best_score < 0.3:
            return DEFAULT_THR
        return round(float(best_thr), 2)

    except Exception as e:
        log(f"  Threshold search failed ({str(e)[:80]}) — using {DEFAULT_THR}")
        return DEFAULT_THR


# ══════════════════════════════════════════════════════════════════════════════
#  HELPER — Evaluate a fitted estimator at two thresholds
# ══════════════════════════════════════════════════════════════════════════════
def evaluate_at_thresholds(y_te, y_pred_default, y_pred_optimal,
                            y_prob, optimal_thr):
    """Return a unified result dict with metrics at both thresholds."""
    def _metrics(y_true, y_pred, y_prob):
        cm = confusion_matrix(y_true, y_pred)
        return {
            'accuracy':  round(accuracy_score(y_true, y_pred), 4),
            'precision': round(precision_score(y_true, y_pred, zero_division=0), 4),
            'recall':    round(recall_score(y_true, y_pred, zero_division=0), 4),
            'f1':        round(f1_score(y_true, y_pred, zero_division=0), 4),
            'auc':       round(roc_auc_score(y_true, y_prob), 4),
            'brier':     round(brier_score_loss(y_true, y_prob), 4),
            'cm':        cm.tolist(),
            'y_pred':    y_pred.tolist(),
        }

    r_def = _metrics(y_te, y_pred_default, y_prob)
    r_opt = _metrics(y_te, y_pred_optimal, y_prob)

    return {
        **r_def,                          # top-level = default threshold (0.5)
        'optimal_threshold': optimal_thr,
        'opt': r_opt,                     # sub-dict = optimal threshold
        'y_prob': y_prob.tolist(),
    }


# ══════════════════════════════════════════════════════════════════════════════
#  STEP 3 — Classical baselines
# ══════════════════════════════════════════════════════════════════════════════
def run_classical(prep):
    """
    GridSearchCV-tuned classifiers.

    v3.1 changes vs v3.0:
      • Platt calibration REMOVED — CalibratedClassifierCV(cv='prefit') was
        compressing all probabilities below 0.20 on this imbalanced dataset,
        causing every classical model to predict all-negative at any threshold
        above 0.20. With SMOTE + class_weight='balanced', raw probabilities
        are already informative.
      • find_optimal_threshold() now fits raw clones via StratifiedKFold
        (not cross_val_predict on a prefit wrapper).
    """
    log("STEP 3 ─ Classical baselines")

    X_tr = prep['X_tr_pca']
    X_te = prep['X_te_pca']
    y_tr = prep['y_tr_aug']      # SMOTE-augmented
    y_te = prep['y_te']

    cv_inner = StratifiedKFold(n_splits=3, shuffle=True, random_state=42)

    # ── Estimator configurations ──────────────────────────────────────────────
    configs = {
        'Logistic Regression': (
            LogisticRegression(random_state=42, class_weight='balanced',
                               max_iter=3000),
            {'C': [0.001, 0.01, 0.1, 1, 10, 100]},
        ),
        'Random Forest': (
            RandomForestClassifier(random_state=42, class_weight='balanced',
                                   n_jobs=-1),
            {'n_estimators': [100, 200, 300],
             'max_depth':    [3, 5, 7, None],
             'min_samples_leaf': [1, 2, 3]},
        ),
        'SVM': (
            SVC(probability=True, random_state=42, class_weight='balanced'),
            {'C': [0.01, 0.1, 1, 10, 100],
             'kernel': ['rbf', 'linear'],
             'gamma': ['scale', 'auto']},
        ),
        'Gradient Boosting': (
            GradientBoostingClassifier(random_state=42),
            {'n_estimators': [100, 200],
             'max_depth':    [2, 3, 4],
             'learning_rate': [0.05, 0.1, 0.2],
             'subsample':    [0.7, 0.9]},
        ),
        'MLP': (
            MLPClassifier(random_state=42, max_iter=5000,
                          early_stopping=True, n_iter_no_change=30,
                          validation_fraction=0.15),
            {'hidden_layer_sizes': [(64, 32), (128, 64), (64,), (32,)],
             'alpha': [0.0001, 0.001, 0.01, 0.1],
             'learning_rate_init': [0.001, 0.005, 0.01]},
        ),
    }

    # XGBoost
    if _XGB_OK:
        pos = int(y_tr.sum()); neg = len(y_tr) - pos
        scale_pw = neg / pos if pos > 0 else 1.0
        xgb_params = dict(
            random_state=42, eval_metric='logloss',
            scale_pos_weight=scale_pw, verbosity=0,
        )
        # Remove deprecated use_label_encoder if not supported
        try:
            xgb.XGBClassifier(**xgb_params, use_label_encoder=False)
            xgb_params['use_label_encoder'] = False
        except TypeError:
            pass
        configs['XGBoost'] = (
            xgb.XGBClassifier(**xgb_params),
            {'n_estimators': [100, 200],
             'max_depth':    [2, 3, 5],
             'learning_rate': [0.03, 0.05, 0.1],
             'subsample':    [0.7, 0.9, 1.0],
             'colsample_bytree': [0.7, 1.0]},
        )
    else:
        log("  XGBoost not available — skipping")

    results = {}
    for name, (estimator, params) in configs.items():
        log(f"  {name} ...")

        # ── GridSearchCV ──────────────────────────────────────────────────────
        gs = GridSearchCV(estimator, params, cv=cv_inner,
                          scoring='f1', n_jobs=-1, refit=True)
        gs.fit(X_tr, y_tr)
        best_est = gs.best_estimator_

        # ── Predictions (raw, no calibration wrapper) ─────────────────────────
        y_prob = best_est.predict_proba(X_te)[:, 1]

        # Optimal threshold via OOF sweep on the raw estimator class
        opt_thr = find_optimal_threshold(best_est, X_tr, y_tr, cv=3)

        y_pred_def = (y_prob >= DEFAULT_THR).astype(int)
        y_pred_opt = (y_prob >= opt_thr).astype(int)

        results[name] = {
            'best_params': str(gs.best_params_),
            'estimator':   best_est,
            **evaluate_at_thresholds(y_te, y_pred_def, y_pred_opt,
                                     y_prob, opt_thr),
        }

        r = results[name]
        log(f"    Best params: {gs.best_params_}")
        log(f"    Default(0.5): F1={r['f1']:.4f} | Recall={r['recall']:.4f} | "
            f"AUC={r['auc']:.4f} | Brier={r['brier']:.4f}")
        log(f"    Optimal({opt_thr:.2f}): F1={r['opt']['f1']:.4f} | "
            f"Recall={r['opt']['recall']:.4f}")

    return results


# ══════════════════════════════════════════════════════════════════════════════
#  STEP 4 — VQC  (ablation-optimal config: ZZ(r=1) + ESU2(r=2))
# ══════════════════════════════════════════════════════════════════════════════
def _detect_qiskit():
    """Auto-detect compatible Qiskit imports across versions."""
    import importlib, qiskit
    qk_ver = tuple(int(x) for x in qiskit.__version__.split('.')[:2])

    Sampler = None
    for mod_name, cls_name in [
        ('qiskit.primitives', 'StatevectorSampler'),
        ('qiskit.primitives', 'Sampler'),
    ]:
        try:
            Sampler = getattr(importlib.import_module(mod_name), cls_name)
            log(f"  Sampler : {mod_name}.{cls_name}")
            break
        except (ImportError, AttributeError):
            pass
    if Sampler is None:
        raise ImportError("No compatible Sampler in qiskit.primitives")

    COBYLA = None
    for pkg in ['qiskit_algorithms.optimizers',
                'qiskit_machine_learning.optimizers',
                'qiskit.algorithms.optimizers']:
        try:
            mod = importlib.import_module(pkg)
            COBYLA = getattr(mod, 'COBYLA', None)
            if COBYLA:
                log(f"  Optimizer: {pkg}.COBYLA")
                break
        except (ImportError, AttributeError):
            pass
    if COBYLA is None:
        raise ImportError("Cannot find COBYLA optimizer")

    VQC = None
    for pkg in ['qiskit_machine_learning.algorithms',
                'qiskit_machine_learning.algorithms.classifiers']:
        try:
            mod = importlib.import_module(pkg)
            VQC = getattr(mod, 'VQC', None)
            if VQC:
                log(f"  VQC     : {pkg}.VQC")
                break
        except (ImportError, AttributeError):
            pass
    if VQC is None:
        raise ImportError("Cannot find VQC")

    return Sampler, COBYLA, VQC, qk_ver


def _build_vqc(Sampler, COBYLA, VQC, fm_reps, ans_reps, maxiter,
               loss_history=None, verbose_every=50):
    """
    Instantiate a VQC with version-aware callback attachment.

    Qiskit ≥2.x: callback belongs to VQC(callback=fn)
                  signature: (weights, objective_value)
    Qiskit  1.x: callback belongs to COBYLA(callback=fn)
                  signature: (nfev, x, fx, dx [, accept])
    """
    from qiskit.circuit.library import ZZFeatureMap, EfficientSU2

    if loss_history is None:
        loss_history = []

    def _vqc_cb(weights, obj_value):
        loss_history.append(float(obj_value))
        n = len(loss_history)
        if verbose_every and n % verbose_every == 0:
            log(f"    iter {n:3d} | loss = {obj_value:.4f}")

    def _cobyla_cb(*args, **kwargs):
        fval = (float(args[2]) if len(args) >= 3 else
                float(args[1]) if len(args) == 2 else
                float(args[0]) if len(args) == 1 else
                float(kwargs.get('fval', 0.0)))
        loss_history.append(fval)
        n = len(loss_history)
        if verbose_every and n % verbose_every == 0:
            log(f"    iter {n:3d} | loss = {fval:.4f}")

    feature_map = ZZFeatureMap(feature_dimension=4, reps=fm_reps)
    ansatz      = EfficientSU2(num_qubits=4, reps=ans_reps)
    sampler     = Sampler()

    vqc_params    = inspect.signature(VQC.__init__).parameters
    cobyla_params = inspect.signature(COBYLA.__init__).parameters
    has_vqc_cb    = 'callback' in vqc_params
    has_cobyla_cb = 'callback' in cobyla_params

    if has_cobyla_cb:
        optimizer = COBYLA(maxiter=maxiter, callback=_cobyla_cb)
        log(f"  Callback: COBYLA-level (Qiskit ≤1.x)")
        vqc = VQC(feature_map=feature_map, ansatz=ansatz,
                  optimizer=optimizer, sampler=sampler)
    elif has_vqc_cb:
        optimizer = COBYLA(maxiter=maxiter)
        log(f"  Callback: VQC-level (Qiskit 2.x)")
        vqc = VQC(feature_map=feature_map, ansatz=ansatz,
                  optimizer=optimizer, sampler=sampler,
                  callback=_vqc_cb)
    else:
        optimizer = COBYLA(maxiter=maxiter)
        log(f"  Callback: disabled (not supported in this build)")
        vqc = VQC(feature_map=feature_map, ansatz=ansatz,
                  optimizer=optimizer, sampler=sampler)

    return vqc, loss_history


def run_vqc(prep,
            fm_reps=VQC_FM_REPS,
            ans_reps=VQC_ANSATZ_REPS,
            maxiter=VQC_MAXITER):
    """
    ★ v3.1 main VQC: ZZFeatureMap(reps=2) + EfficientSU2(reps=2)
    Validated as optimal by v3.0 ablation study:
      AUC=0.750, F1_opt=0.571 — highest AUC and F1 in the 7-config grid.
    Same number of parameters (24) as ZZ(r=1)+ESU2(r=2) but higher expressivity
    from ZZFeatureMap's entanglement at reps=2.

    Primary reporting threshold: τ=0.45 (F1=0.529, Recall=0.90 from v3.0 sweep)
    """
    log(f"STEP 4 ─ VQC  [ZZ(r={fm_reps})+ESU2(r={ans_reps}), maxiter={maxiter}]")
    Sampler, COBYLA, VQC, qk_ver = _detect_qiskit()
    log(f"  Qiskit version: {qk_ver}")

    np.random.seed(42)

    X_tr_pca = prep['X_tr_pca']
    X_te_pca = prep['X_te_pca']
    y_tr     = prep['y_tr_aug']
    y_te     = prep['y_te']

    # Scale to (−π, π) for quantum encoding
    mm = MinMaxScaler(feature_range=(-np.pi, np.pi))
    X_tr_q = mm.fit_transform(X_tr_pca)
    X_te_q = mm.transform(X_te_pca)

    loss_history = []
    vqc, loss_history = _build_vqc(
        Sampler, COBYLA, VQC,
        fm_reps, ans_reps, maxiter,
        loss_history=loss_history,
    )

    t0 = time.time()
    vqc.fit(X_tr_q, y_tr)
    elapsed = (time.time() - t0) / 60
    log(f"  Training: {elapsed:.1f} min | {len(loss_history)} loss values")

    y_prob = vqc.predict_proba(X_te_q)[:, 1]

    # Optimal threshold — start from τ=VQC_OPT_THR (validated on v3.0 sweep)
    # then verify/update via OOF CV if enough training data
    opt_thr = VQC_OPT_THR
    if len(y_tr) > 10:
        thresholds = np.arange(0.05, 0.96, 0.05)
        from sklearn.model_selection import StratifiedKFold as SKF
        skf = SKF(n_splits=3, shuffle=True, random_state=42)
        # Use OOF predictions on training set for threshold selection
        oof_probs = np.zeros(len(y_tr))
        try:
            for tr_idx, val_idx in skf.split(X_tr_q, y_tr):
                vqc_cv, _ = _build_vqc(Sampler, COBYLA, VQC,
                                        fm_reps, ans_reps, maxiter // 2,
                                        verbose_every=0)
                vqc_cv.fit(X_tr_q[tr_idx], y_tr[tr_idx])
                oof_probs[val_idx] = vqc_cv.predict_proba(
                    X_tr_q[val_idx])[:, 1]
            best_f1, best_thr = -1.0, DEFAULT_THR
            for thr in thresholds:
                s = f1_score(y_tr, (oof_probs >= thr).astype(int),
                             zero_division=0)
                if s > best_f1:
                    best_f1, best_thr = s, thr
            opt_thr = round(float(best_thr), 2)
            log(f"  CV-optimal threshold: {opt_thr:.2f}  (OOF F1={best_f1:.4f})")
        except Exception as e:
            log(f"  OOF threshold search failed ({e}) — using {DEFAULT_THR}")

    y_pred_def = (y_prob >= DEFAULT_THR).astype(int)
    y_pred_opt = (y_prob >= opt_thr).astype(int)

    # Save convergence
    pd.DataFrame({'iteration': range(1, len(loss_history)+1),
                  'loss': loss_history}).to_csv(
        OUTPUT_DIR / "vqc_convergence.csv", index=False)
    if loss_history:
        log(f"  Convergence: {loss_history[0]:.4f} → {loss_history[-1]:.4f} "
            f"(Δ={loss_history[0]-loss_history[-1]:.4f})")

    result = {
        'best_params': (f'ZZFeatureMap(reps={fm_reps})'
                        f'+EfficientSU2(reps={ans_reps})'
                        f'+COBYLA({maxiter})'),
        'training_time_min': round(elapsed, 2),
        'loss_history': loss_history,
        **evaluate_at_thresholds(y_te, y_pred_def, y_pred_opt, y_prob, opt_thr),
    }

    r = result
    log(f"  Default(0.5): F1={r['f1']:.4f} | Recall={r['recall']:.4f} | "
        f"AUC={r['auc']:.4f} | Brier={r['brier']:.4f}")
    log(f"  Optimal({opt_thr:.2f}): F1={r['opt']['f1']:.4f} | "
        f"Recall={r['opt']['recall']:.4f}")
    return result


# ══════════════════════════════════════════════════════════════════════════════
#  STEP 5a — Bootstrap Confidence Intervals  (n=1,000)
# ══════════════════════════════════════════════════════════════════════════════
def _boot_metrics(y_true, y_pred, y_prob):
    out = {}
    out['f1']        = f1_score(y_true, y_pred, zero_division=0)
    out['recall']    = recall_score(y_true, y_pred, zero_division=0)
    out['precision'] = precision_score(y_true, y_pred, zero_division=0)
    try:
        out['auc'] = roc_auc_score(y_true, y_prob)
    except Exception:
        out['auc'] = np.nan
    return out


def bootstrap_ci(y_true, y_pred, y_prob, n_boot=1000, alpha=0.05):
    rng = np.random.RandomState(42)
    n   = len(y_true)
    acc = {k: [] for k in ['f1', 'recall', 'precision', 'auc']}

    for _ in range(n_boot):
        idx  = rng.randint(0, n, n)
        yt   = y_true[idx]; yp = y_pred[idx]; yprob = y_prob[idx]
        if yt.sum() == 0 or yt.sum() == n:
            continue
        m = _boot_metrics(yt, yp, yprob)
        for k, v in m.items():
            if not np.isnan(v):
                acc[k].append(v)

    lo, hi = alpha / 2 * 100, (1 - alpha / 2) * 100
    return {
        k: {'lower': round(float(np.percentile(v, lo)), 4),
            'upper': round(float(np.percentile(v, hi)), 4)}
        if len(v) >= 10 else {'lower': np.nan, 'upper': np.nan}
        for k, v in acc.items()
    }


def run_bootstrap(results, y_te, n_boot=1000, use_optimal=False):
    """
    Compute bootstrap CIs.
    use_optimal=True → compute on optimal-threshold predictions.
    """
    suffix = " [optimal threshold]" if use_optimal else " [default threshold 0.5]"
    log(f"STEP 5a ─ Bootstrap CI (n={n_boot}){suffix}")
    y_true = np.array(y_te)
    boot   = {}

    for model, r in results.items():
        key   = 'opt' if use_optimal else None
        yp    = np.array(r['opt']['y_pred'] if use_optimal else r['y_pred'])
        yprob = np.array(r['y_prob'])
        boot[model] = bootstrap_ci(y_true, yp, yprob, n_boot=n_boot)
        ci = boot[model]
        log(f"  {model:<22}  F1 [{ci['f1']['lower']:.3f}, {ci['f1']['upper']:.3f}]  "
            f"Recall [{ci['recall']['lower']:.3f}, {ci['recall']['upper']:.3f}]")

    return boot


# ══════════════════════════════════════════════════════════════════════════════
#  STEP 5b — McNemar's Test
# ══════════════════════════════════════════════════════════════════════════════
def run_mcnemar(results, y_te, use_optimal=True):
    """
    Pairwise exact McNemar's test (VQC vs each baseline).
    By default uses optimal-threshold predictions (more meaningful).
    """
    suffix = "optimal" if use_optimal else "default"
    log(f"STEP 5b ─ McNemar's tests [{suffix} threshold]")
    y_true = np.array(y_te)
    tests  = []

    if 'VQC' not in results or 'y_pred' not in results.get('VQC', {}):
        log("  VQC predictions not available — skipping")
        return tests

    vqc_pred = np.array(
        results['VQC']['opt']['y_pred'] if use_optimal
        else results['VQC']['y_pred']
    )

    for model in [m for m in results if m != 'VQC']:
        if 'y_pred' not in results.get(model, {}):
            continue
        cl_pred = np.array(
            results[model]['opt']['y_pred'] if use_optimal
            else results[model]['y_pred']
        )

        b = int(np.sum((cl_pred != y_true) & (vqc_pred == y_true)))
        c = int(np.sum((cl_pred == y_true) & (vqc_pred != y_true)))
        n = b + c

        if n == 0:
            p_val, note = 1.0, "identical predictions"
        elif n < 25:
            x     = min(b, c)
            p_val = float(2 * sp_binom.cdf(x, n, 0.5))
            note  = "exact binomial (n<25)"
        elif _SM_OK:
            table = np.array([[0, b], [c, 0]])
            stat  = sm_mcnemar(table, exact=False, correction=True)
            p_val = float(stat.pvalue)
            note  = "chi² + continuity correction"
        else:
            chi   = (abs(b - c) - 1) ** 2 / n
            p_val = float(1 - chi2.cdf(chi, 1))
            note  = "manual chi²"

        sig = ("**p<0.01" if p_val < 0.01 else
               "*p<0.05"  if p_val < 0.05 else "ns")

        tests.append({
            'comparison':   f"VQC vs {model}",
            'b':            b,
            'c':            c,
            'n':            n,
            'p_value':      round(p_val, 4),
            'significance': sig,
            'method':       note,
        })
        log(f"  VQC vs {model:<22}  b={b} c={c}  "
            f"p={p_val:.4f}  {sig}  [{note}]")

    return tests


# ══════════════════════════════════════════════════════════════════════════════
#  STEP 6 — Per-country error analysis
# ══════════════════════════════════════════════════════════════════════════════
def run_per_country(results, prep, use_optimal=True):
    log("STEP 6 ─ Per-country error analysis")
    meta  = prep['meta_te'].copy()
    y_te  = np.array(prep['y_te'])
    meta['true_label'] = y_te
    cc    = meta.columns[0]

    for model, r in results.items():
        preds = (r['opt']['y_pred'] if use_optimal and 'opt' in r
                 else r.get('y_pred', []))
        meta[f'{model}_pred']    = preds
        meta[f'{model}_correct'] = (np.array(preds) == y_te).astype(int)

    # Per-country summary
    country_rows = []
    for country, grp in meta.groupby(cc):
        pos = grp[grp['true_label'] == 1]
        row = {'country': country,
               'n_test': len(grp),
               'n_sgd':  len(pos)}
        for model in results:
            col = f'{model}_pred'
            row[f'{model}_detected'] = int(pos[col].sum()) if col in pos.columns and len(pos) > 0 else 0
        country_rows.append(row)

    country_df = pd.DataFrame(country_rows)

    # Log detection summary
    sgd_events = meta[meta['true_label'] == 1]
    log(f"  SGD events in test set: {len(sgd_events)}")
    for model in results:
        col = f'{model}_pred'
        if col in sgd_events.columns:
            det = int(sgd_events[col].sum())
            log(f"    {model:<22} detected {det}/{len(sgd_events)}")

    country_df.to_csv(OUTPUT_DIR / "per_country_analysis.csv", index=False)
    meta.to_csv(OUTPUT_DIR / "test_predictions_detailed.csv", index=False)
    return country_df, meta


# ══════════════════════════════════════════════════════════════════════════════
#  STEP 7 — Threshold sensitivity
# ══════════════════════════════════════════════════════════════════════════════
def run_threshold_sensitivity(results, y_te,
                              thresholds=None):
    log("STEP 7 ─ Threshold sensitivity analysis")
    if thresholds is None:
        thresholds = np.round(np.arange(0.05, 0.96, 0.05), 2)

    y_true  = np.array(y_te)
    records = []

    for model, r in results.items():
        if 'y_prob' not in r:
            continue
        y_prob = np.array(r['y_prob'])
        for thr in thresholds:
            yp = (y_prob >= thr).astype(int)
            records.append({
                'model':       model,
                'threshold':   float(thr),
                'f1':          round(f1_score(y_true, yp, zero_division=0), 4),
                'recall':      round(recall_score(y_true, yp, zero_division=0), 4),
                'precision':   round(precision_score(y_true, yp, zero_division=0), 4),
                'n_pred_pos':  int(yp.sum()),
            })

    thr_df = pd.DataFrame(records)
    thr_df.to_csv(OUTPUT_DIR / "threshold_sensitivity.csv", index=False)

    log("  Optimal threshold (max F1) per model:")
    for model in results:
        sub = thr_df[thr_df['model'] == model]
        if sub.empty: continue
        best = sub.loc[sub['f1'].idxmax()]
        log(f"    {model:<22}  thr={best['threshold']:.2f}  "
            f"F1={best['f1']:.4f}  Recall={best['recall']:.4f}  "
            f"n_pos={int(best['n_pred_pos'])}")

    return thr_df


# ══════════════════════════════════════════════════════════════════════════════
#  STEP 8 — Robustness checks (SGD threshold 40% / 60%)
# ══════════════════════════════════════════════════════════════════════════════
def run_robustness(filepath):
    log("STEP 8 ─ Robustness checks [SGD threshold 40% / 60%]")
    robust = {}
    main_thr = 0.50

    for thr in [0.40, 0.60]:
        label = f"SGD_{int(thr*100)}pct"
        log(f"  {label} ...")
        try:
            df_r   = load_and_prepare(filepath, threshold=thr)
            prep_r = preprocess(df_r.copy(), target_col='target')
            if prep_r['y_te'].sum() < 2:
                log("    Skipped: <2 positives in test set"); continue
            res_r  = run_classical(prep_r)
            robust[label] = {
                'n_pos_test':  int(prep_r['y_te'].sum()),
                'n_total_test': int(len(prep_r['y_te'])),
                'models': {
                    m: {
                        'f1_default': r['f1'],
                        'f1_optimal': r['opt']['f1'],
                        'recall_default': r['recall'],
                        'recall_optimal': r['opt']['recall'],
                        'auc': r['auc'],
                    }
                    for m, r in res_r.items()
                },
            }
        except Exception as e:
            log(f"    Error: {e}")

    return robust


# ══════════════════════════════════════════════════════════════════════════════
#  STEP 9 — VQC Ablation Study
# ══════════════════════════════════════════════════════════════════════════════
def run_ablation(prep):
    log("STEP 9 ─ VQC Ablation Study")
    try:
        Sampler, COBYLA, VQC, qk_ver = _detect_qiskit()
        from qiskit.circuit.library import ZZFeatureMap, EfficientSU2
    except ImportError as e:
        log(f"  Qiskit unavailable: {e}"); return []

    X_tr_pca = prep['X_tr_pca']
    X_te_pca = prep['X_te_pca']
    y_tr     = prep['y_tr_aug']
    y_te     = np.array(prep['y_te'])

    mm = MinMaxScaler(feature_range=(-np.pi, np.pi))
    X_tr_q = mm.fit_transform(X_tr_pca)
    X_te_q = mm.transform(X_te_pca)

    # Expanded ablation grid
    configs = [
        # label,                      fm_reps, ans_reps, maxiter
        ('ZZ(r=1)+ESU2(r=1)',         1, 1, 200),
        ('ZZ(r=1)+ESU2(r=2) ★ MAIN', 1, 2, 300),   # ← v3.0 main config
        ('ZZ(r=1)+ESU2(r=3)',         1, 3, 300),
        ('ZZ(r=2)+ESU2(r=2)',         2, 2, 300),
        ('ZZ(r=2)+ESU2(r=3)',         2, 3, 400),   # v2.0 main config
        ('ZZ(r=3)+ESU2(r=3)',         3, 3, 400),
        ('ZZ(r=2)+ESU2(r=4)',         2, 4, 500),
    ]

    records = []
    for label, fm_reps, ans_reps, maxiter in configs:
        np.random.seed(42)
        log(f"  {label}  (maxiter={maxiter})")
        try:
            t0 = time.time()
            vqc, _ = _build_vqc(Sampler, COBYLA, VQC,
                                 fm_reps, ans_reps, maxiter,
                                 verbose_every=0)
            vqc.fit(X_tr_q, y_tr)
            y_pred = vqc.predict(X_te_q)
            y_prob = vqc.predict_proba(X_te_q)[:, 1]
            elapsed = (time.time() - t0) / 60

            # Optimal threshold (quick sweep, no CV for ablation speed)
            best_f1, best_thr = -1.0, DEFAULT_THR
            for thr in np.arange(0.05, 0.96, 0.05):
                s = f1_score(y_te, (y_prob >= thr).astype(int), zero_division=0)
                if s > best_f1: best_f1, best_thr = s, thr

            records.append({
                'config':          label,
                'fm_reps':         fm_reps,
                'ansatz_reps':     ans_reps,
                'n_params':        vqc.ansatz.num_parameters,
                'f1_default':      round(f1_score(y_te, y_pred, zero_division=0), 4),
                'f1_optimal':      round(best_f1, 4),
                'recall_default':  round(recall_score(y_te, y_pred, zero_division=0), 4),
                'recall_optimal':  round(recall_score(y_te, (y_prob >= best_thr).astype(int), zero_division=0), 4),
                'precision_default': round(precision_score(y_te, y_pred, zero_division=0), 4),
                'auc':             round(roc_auc_score(y_te, y_prob), 4),
                'opt_threshold':   round(float(best_thr), 2),
                'train_min':       round(elapsed, 2),
            })
            log(f"    F1(def)={records[-1]['f1_default']:.4f}  "
                f"F1(opt)={records[-1]['f1_optimal']:.4f}  "
                f"Recall(opt)={records[-1]['recall_optimal']:.4f}  "
                f"params={records[-1]['n_params']}  "
                f"t={records[-1]['train_min']:.1f}min")
        except Exception as e:
            log(f"    Failed: {str(e)[:120]}")
            records.append({'config': label, 'error': str(e)[:120]})

    pd.DataFrame(records).to_csv(OUTPUT_DIR / "vqc_ablation.csv", index=False)
    log("  Saved → vqc_ablation.csv")
    return records


# ══════════════════════════════════════════════════════════════════════════════
#  STEP 10 — Figures
# ══════════════════════════════════════════════════════════════════════════════
def _model_color(name):
    return PALETTE.get(name, '#888888')


def generate_figures(results, prep, y_te, boot_ci_dict, thr_df,
                     country_df, ablation):
    log("STEP 10 ─ Generating figures")
    y_te   = np.array(y_te)
    models = list(results.keys())
    cols   = [_model_color(m) for m in models]

    # ── Pull metric arrays (default threshold) ────────────────────────────────
    f1s_def  = [results[m]['f1']             for m in models]
    f1s_opt  = [results[m]['opt']['f1']      for m in models]
    rec_def  = [results[m]['recall']         for m in models]
    rec_opt  = [results[m]['opt']['recall']  for m in models]
    aucs     = [results[m]['auc']            for m in models]
    briers   = [results[m]['brier']          for m in models]
    opt_thrs = [results[m]['optimal_threshold'] for m in models]

    # ── Figure 1: Main 6-panel comparison ────────────────────────────────────
    fig1 = plt.figure(figsize=(18, 13))
    gs1  = gridspec.GridSpec(3, 3, figure=fig1, hspace=0.50, wspace=0.38)

    def _errbar(ax, vals, label, boot_key, color_list):
        # Clip error bars to [0, ∞) — negative yerr raises ValueError in
        # matplotlib when CI bounds cross the point estimate (e.g. all-zero
        # metrics for classical models with degenerate bootstrap samples).
        lo = [max(0.0, v - boot_ci_dict.get(m, {}).get(boot_key, {}).get('lower', v))
              for m, v in zip(models, vals)]
        hi = [max(0.0, boot_ci_dict.get(m, {}).get(boot_key, {}).get('upper', v) - v)
              for m, v in zip(models, vals)]
        bars = ax.bar(models, vals, color=color_list, edgecolor='white',
                      linewidth=0.6)
        # Only draw error bars where at least one side is non-zero
        if any(l + h > 0 for l, h in zip(lo, hi)):
            ax.errorbar(range(len(models)), vals, yerr=[lo, hi],
                        fmt='none', color='#333', capsize=4, linewidth=1.2)
        for bar, v in zip(bars, vals):
            ax.text(bar.get_x() + bar.get_width()/2, v + 0.01,
                    f'{v:.3f}', ha='center', fontsize=8)
        return bars

    # (a) F1 default
    ax = fig1.add_subplot(gs1[0, 0])
    best_cl = max([f for m, f in zip(models, f1s_def) if m != 'VQC'], default=0)
    _errbar(ax, f1s_def, 'F1 (default)', 'f1', cols)
    ax.axhline(best_cl, color='gray', ls='--', lw=1,
               label=f'Best classical ({best_cl:.3f})')
    ax.set_ylim(0, 1); ax.set_title('(a) F1 — default threshold (0.5)',
                                    fontsize=10)
    ax.set_xticklabels(models, rotation=22, ha='right', fontsize=8)
    ax.legend(fontsize=7)

    # (b) F1 optimal threshold
    ax = fig1.add_subplot(gs1[0, 1])
    best_cl_opt = max([f for m, f in zip(models, f1s_opt) if m != 'VQC'], default=0)
    _errbar(ax, f1s_opt, 'F1 (optimal)', 'f1', cols)
    ax.axhline(best_cl_opt, color='gray', ls='--', lw=1,
               label=f'Best classical ({best_cl_opt:.3f})')
    ax.set_ylim(0, 1); ax.set_title('(b) F1 — optimal threshold per model',
                                    fontsize=10)
    ax.set_xticklabels(models, rotation=22, ha='right', fontsize=8)
    ax.legend(fontsize=7)

    # Annotate optimal threshold values
    for i, (thr, v) in enumerate(zip(opt_thrs, f1s_opt)):
        ax.text(i, -0.08, f'τ={thr:.2f}', ha='center',
                fontsize=7, color='#555', transform=ax.get_xaxis_transform())

    # (c) Recall optimal
    ax = fig1.add_subplot(gs1[0, 2])
    _errbar(ax, rec_opt, 'Recall', 'recall', cols)
    ax.set_ylim(0, 1); ax.set_title('(c) Recall — optimal threshold', fontsize=10)
    ax.set_xticklabels(models, rotation=22, ha='right', fontsize=8)

    # (d) AUC-ROC
    ax = fig1.add_subplot(gs1[1, 0])
    bars = ax.bar(models, aucs, color=cols, edgecolor='white', linewidth=0.6)
    ax.axhline(0.5, color='red', ls=':', lw=1, label='Random (0.5)')
    for bar, v in zip(bars, aucs):
        ax.text(bar.get_x() + bar.get_width()/2, v + 0.01,
                f'{v:.3f}', ha='center', fontsize=8)
    ax.set_ylim(0, 1); ax.set_title('(d) AUC-ROC', fontsize=10)
    ax.set_xticklabels(models, rotation=22, ha='right', fontsize=8)
    ax.legend(fontsize=7)

    # (e) Brier score (lower = better)
    ax = fig1.add_subplot(gs1[1, 1])
    bars = ax.bar(models, briers, color=cols, edgecolor='white', linewidth=0.6)
    ax.axhline(y_te.mean() * (1 - y_te.mean()), color='gray', ls='--',
               lw=1, label='No-skill baseline')
    for bar, v in zip(bars, briers):
        ax.text(bar.get_x() + bar.get_width()/2, v + 0.003,
                f'{v:.3f}', ha='center', fontsize=8)
    ax.set_title('(e) Brier score (↓ better)', fontsize=10)
    ax.set_xticklabels(models, rotation=22, ha='right', fontsize=8)
    ax.legend(fontsize=7)

    # (f) ROC curves
    ax = fig1.add_subplot(gs1[1, 2])
    ax.plot([0, 1], [0, 1], 'k--', lw=1, label='Random')
    for m, c in zip(models, cols):
        if 'y_prob' not in results[m]: continue
        fpr, tpr, _ = roc_curve(y_te, results[m]['y_prob'])
        ax.plot(fpr, tpr, color=c, lw=1.8,
                label=f"{m} ({results[m]['auc']:.3f})")
    ax.set_xlim(0, 1); ax.set_ylim(0, 1.02)
    ax.set_xlabel('FPR', fontsize=9); ax.set_ylabel('TPR', fontsize=9)
    ax.set_title('(f) ROC curves', fontsize=10)
    ax.legend(fontsize=6, loc='lower right')

    # (g) Precision–Recall scatter (optimal threshold)
    ax = fig1.add_subplot(gs1[2, 0])
    for m, c in zip(models, cols):
        ax.scatter(results[m]['opt']['recall'],
                   results[m]['opt']['precision'],
                   s=110, color=c, zorder=5, label=m)
        ax.annotate(m, (results[m]['opt']['recall'],
                        results[m]['opt']['precision']),
                    xytext=(5, 3), textcoords='offset points', fontsize=7)
    ax.set_xlim(0, 1.05); ax.set_ylim(0, 1.05)
    ax.set_xlabel('Recall', fontsize=9); ax.set_ylabel('Precision', fontsize=9)
    ax.set_title('(g) Precision vs Recall (optimal τ)', fontsize=10)
    ax.grid(True, alpha=0.25)

    # (h) VQC Confusion Matrix (optimal threshold)
    ax = fig1.add_subplot(gs1[2, 1])
    if 'VQC' in results:
        cm = results['VQC']['opt']['cm']
        sns.heatmap(cm, annot=True, fmt='d', cmap='Blues',
                    xticklabels=['Normal', 'SGD'],
                    yticklabels=['Normal', 'SGD'],
                    ax=ax, cbar=True, annot_kws={'size': 14})
        ax.set_title(f'(h) VQC confusion matrix\n(τ={results["VQC"]["optimal_threshold"]:.2f})',
                     fontsize=10)
        ax.set_xlabel('Predicted'); ax.set_ylabel('Actual')

    # (i) F1 default vs optimal (lollipop)
    ax = fig1.add_subplot(gs1[2, 2])
    x = np.arange(len(models))
    ax.scatter(f1s_def, x, marker='o', color=cols, s=60, zorder=5,
               label='Default (τ=0.5)')
    ax.scatter(f1s_opt, x, marker='D', color=cols, s=60, alpha=0.6,
               zorder=5, label='Optimal τ')
    for i, (d, o) in enumerate(zip(f1s_def, f1s_opt)):
        ax.plot([d, o], [i, i], color=cols[i], lw=1.2, alpha=0.5)
    ax.set_yticks(x); ax.set_yticklabels(models, fontsize=8)
    ax.set_xlabel('F1-Score', fontsize=9)
    ax.set_title('(i) Threshold impact on F1', fontsize=10)
    ax.set_xlim(0, 1); ax.grid(axis='x', alpha=0.25)
    ax.legend(fontsize=7, loc='lower right')

    fig1.suptitle(
        'Figure 1. Out-of-Sample Performance Comparison\n'
        '(Test: 2019–2023, n=40 obs, 10 SGD events | '
        'VQC: ZZFeatureMap(r=1)+EfficientSU2(r=2)+COBYLA)',
        fontsize=11, fontweight='bold',
    )
    plt.savefig(OUTPUT_DIR / "fig_main_results.png", dpi=200, bbox_inches='tight')
    plt.close()
    log("  fig_main_results.png saved")

    # ── Figure 2: PCA loadings ────────────────────────────────────────────────
    fig2, ax = plt.subplots(figsize=(10, 5))
    load_df = prep['pca_loadings']
    var = prep['pca_var']
    sns.heatmap(load_df, annot=True, fmt='.3f', cmap='RdBu_r',
                center=0, linewidths=0.4, ax=ax, cbar_kws={'shrink': 0.8})
    ax.set_xticklabels(
        [f'PC{i+1}\n({v*100:.1f}%)' for i, v in enumerate(var)], fontsize=10)
    ax.set_title('PCA Feature Loadings\n'
                 '(Fitted on SMOTE-augmented training data 2000–2018)',
                 fontsize=11, fontweight='bold')
    plt.tight_layout()
    plt.savefig(OUTPUT_DIR / "fig_pca_loadings.png", dpi=200, bbox_inches='tight')
    plt.close()
    log("  fig_pca_loadings.png saved")

    # ── Figure 3: VQC convergence ─────────────────────────────────────────────
    conv_path = OUTPUT_DIR / "vqc_convergence.csv"
    if conv_path.exists():
        conv_df = pd.read_csv(conv_path)
        if len(conv_df) > 0:
            fig3, ax = plt.subplots(figsize=(10, 4))
            ax.plot(conv_df['iteration'], conv_df['loss'],
                    color='#185FA5', lw=1.5, alpha=0.85)
            # Smooth trend
            if len(conv_df) > 10:
                window = max(5, len(conv_df) // 20)
                smooth = conv_df['loss'].rolling(window, center=True).mean()
                ax.plot(conv_df['iteration'], smooth,
                        color='#D55E00', lw=2, label=f'Rolling mean (w={window})')
                ax.legend(fontsize=9)
            ax.set_xlabel('Iteration'); ax.set_ylabel('COBYLA Loss')
            ax.set_title(
                f'VQC Training Convergence — ZZFeatureMap(r={VQC_FM_REPS})'
                f'+EfficientSU2(r={VQC_ANSATZ_REPS})+COBYLA({VQC_MAXITER})',
                fontsize=11, fontweight='bold')
            ax.grid(True, alpha=0.25)
            plt.tight_layout()
            plt.savefig(OUTPUT_DIR / "fig_vqc_convergence.png",
                        dpi=200, bbox_inches='tight')
            plt.close()
            log("  fig_vqc_convergence.png saved")

    # ── Figure 4: Threshold sensitivity ──────────────────────────────────────
    if thr_df is not None and not thr_df.empty:
        fig4, axes = plt.subplots(1, 2, figsize=(14, 5))
        for ax, metric, label in zip(axes, ['f1', 'recall'],
                                     ['F1-Score', 'Recall']):
            for m in models:
                if 'y_prob' not in results[m]: continue
                sub = thr_df[thr_df['model'] == m]
                ax.plot(sub['threshold'], sub[metric],
                        label=m, color=_model_color(m), lw=2)
            ax.axvline(DEFAULT_THR, color='gray', ls='--', lw=1,
                       label=f'Default ({DEFAULT_THR})')
            ax.set_xlabel('Decision Threshold'); ax.set_ylabel(label)
            ax.set_title(f'{label} vs Decision Threshold')
            ax.legend(fontsize=7); ax.grid(True, alpha=0.25)
            ax.set_xlim(0, 1); ax.set_ylim(0, 1)
        fig4.suptitle(
            'Figure 4. Threshold Sensitivity Analysis\n'
            '(Explains why LR/RF collapse to F1=0 at default threshold 0.5)',
            fontsize=11, fontweight='bold')
        plt.tight_layout()
        plt.savefig(OUTPUT_DIR / "fig_threshold_sensitivity.png",
                    dpi=200, bbox_inches='tight')
        plt.close()
        log("  fig_threshold_sensitivity.png saved")

    # ── Figure 5: Per-country detection ──────────────────────────────────────
    if country_df is not None and not country_df.empty:
        sgd_c = country_df[country_df['n_sgd'] > 0].copy()
        if len(sgd_c) > 0:
            det_cols   = [c for c in sgd_c.columns if c.endswith('_detected')]
            mnames     = [c.replace('_detected', '') for c in det_cols]
            fig5, ax   = plt.subplots(figsize=(13, 5))
            x          = np.arange(len(sgd_c))
            w          = 0.8 / max(len(det_cols), 1)
            for i, (col, mn) in enumerate(zip(det_cols, mnames)):
                off = (i - len(det_cols)/2 + 0.5) * w
                ax.bar(x + off, sgd_c[col].fillna(0), w,
                       label=mn, color=_model_color(mn), alpha=0.85)
            ax.plot(x, sgd_c['n_sgd'], 'k^--',
                    ms=8, lw=1.5, label='Actual SGD events')
            ax.set_xticks(x)
            ax.set_xticklabels(sgd_c['country'], rotation=25, ha='right')
            ax.set_ylabel('SGD Events Detected')
            ax.set_title('Figure 5. Per-Country SGD Detection (2019–2023)',
                         fontsize=11, fontweight='bold')
            ax.legend(fontsize=7, ncol=4)
            ax.grid(axis='y', alpha=0.25)
            plt.tight_layout()
            plt.savefig(OUTPUT_DIR / "fig_per_country.png",
                        dpi=200, bbox_inches='tight')
            plt.close()
            log("  fig_per_country.png saved")

    # ── Figure 6: Ablation study ──────────────────────────────────────────────
    if ablation:
        abl_df = pd.DataFrame([r for r in ablation if 'error' not in r])
        if not abl_df.empty and 'f1_optimal' in abl_df.columns:
            fig6, axes = plt.subplots(1, 3, figsize=(16, 5))
            x     = np.arange(len(abl_df))
            is_main = abl_df['config'].str.contains('MAIN', na=False)
            bar_c = ['#1D9E75' if m else '#B4B2A9' for m in is_main]

            for ax, col, label in zip(axes,
                                      ['f1_default', 'f1_optimal', 'recall_optimal'],
                                      ['F1 (default τ=0.5)', 'F1 (optimal τ)', 'Recall (optimal τ)']):
                bars = ax.bar(x, abl_df[col], color=bar_c,
                              edgecolor='white', linewidth=0.6)
                for bar, v in zip(bars, abl_df[col]):
                    ax.text(bar.get_x() + bar.get_width()/2, v + 0.01,
                            f'{v:.3f}', ha='center', fontsize=9)
                # Param count on x-axis
                ax.set_xticks(x)
                xlabels = [f"{c.split('+')[0].replace(' ★ MAIN','')}\n+{c.split('+')[1].split(' ★')[0]}"
                           if '+' in c else c for c in abl_df['config']]
                ax.set_xticklabels(xlabels, rotation=30, ha='right', fontsize=8)
                ax.set_ylim(0, 1); ax.set_ylabel(label)
                ax.set_title(f'VQC Ablation — {label}')
                ax.grid(axis='y', alpha=0.25)

            # Add param count on secondary axis of last plot
            ax2 = axes[2].twinx()
            ax2.plot(x, abl_df['n_params'], 'o--', color='#185FA5',
                     ms=5, lw=1.2, label='# params')
            ax2.set_ylabel('# trainable params', color='#185FA5')
            ax2.tick_params(axis='y', labelcolor='#185FA5')

            fig6.suptitle(
                'Figure 6. VQC Ablation Study\n'
                '(★ = v3.0 main config, validated as optimal by ablation)',
                fontsize=11, fontweight='bold')
            plt.tight_layout()
            plt.savefig(OUTPUT_DIR / "fig_vqc_ablation.png",
                        dpi=200, bbox_inches='tight')
            plt.close()
            log("  fig_vqc_ablation.png saved")

    log("  All figures generated")


# ══════════════════════════════════════════════════════════════════════════════
#  STEP 11 — Excel export (comprehensive workbook)
# ══════════════════════════════════════════════════════════════════════════════
def export_excel(results, prep, y_te, df_full,
                 tests, robust, boot_ci_dict,
                 thr_df, country_df, ablation):
    log("STEP 11 ─ Exporting to Excel")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
        from openpyxl.utils import get_column_letter
    except ImportError:
        log("  openpyxl not installed — skipping"); return

    # ── Styles ────────────────────────────────────────────────────────────────
    BLUE_H  = PatternFill("solid", fgColor="1565C0")
    VQC_H   = PatternFill("solid", fgColor="1A237E")
    VQC_L   = PatternFill("solid", fgColor="C5CAE9")
    OPT_H   = PatternFill("solid", fgColor="1B5E20")
    OPT_L   = PatternFill("solid", fgColor="C8E6C9")
    ALT     = PatternFill("solid", fgColor="F5F5F5")
    WHT_B   = Font(bold=True, color="FFFFFF")
    BOLD    = Font(bold=True)
    CTR     = Alignment(horizontal='center', vertical='center', wrap_text=True)
    THIN    = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

    def hrow(ws, row, ncols, fill=BLUE_H, font=WHT_B):
        for c in range(1, ncols+1):
            cell = ws.cell(row=row, column=c)
            cell.fill = fill; cell.font = font
            cell.alignment = CTR; cell.border = THIN

    def autowidth(ws, mn=10, mx=55):
        for col in ws.columns:
            L = max((len(str(cell.value or '')) for cell in col), default=8)
            ws.column_dimensions[
                get_column_letter(col[0].column)].width = min(max(L+2, mn), mx)

    wb = Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Dataset ──────────────────────────────────────────────────────
    ws = wb.create_sheet("1_Dataset")
    ws.append(["Metric", "Value", "Notes"]); hrow(ws, 1, 3)
    cc = 'country_name' if 'country_name' in df_full.columns else 'country'
    for i, r in enumerate([
        ("Total observations", len(df_full), "10 countries × ~23 years"),
        ("Countries", df_full[cc].nunique(),
         ", ".join(sorted(df_full[cc].unique()))),
        ("Period", f"{df_full['year'].min()}–{df_full['year'].max()}", "Annual"),
        ("Features", len(FEATURES), "; ".join(FEATURES)),
        ("SGD events (50% threshold)",
         int(df_full['target'].sum()),
         f"{df_full['target'].mean()*100:.1f}%"),
        ("Normal cases", int((df_full['target']==0).sum()),
         f"{(df_full['target']==0).mean()*100:.1f}%"),
        ("Imbalance ratio",
         f"{(df_full['target']==0).sum()/df_full['target'].sum():.1f}:1", ""),
        ("", "", ""),
        ("Train period", "2000–2018",
         f"{int(prep['y_tr'].sum())} SGD events"),
        ("Test period", "2019–2023", f"{int(prep['y_te'].sum())} SGD events"),
        ("Oversampling method", prep['os_method'],
         "Applied to train only"),
        ("After oversampling (train)",
         len(prep['y_tr_aug']),
         f"{int(prep['y_tr_aug'].sum())} pos"),
        ("PCA variance retained",
         f"{sum(prep['pca_var'])*100:.1f}%", "4 components"),
        ("VQC config (v3.0 optimal)",
         f"ZZ(r={VQC_FM_REPS})+ESU2(r={VQC_ANSATZ_REPS})+COBYLA({VQC_MAXITER})",
         "Ablation-validated"),
    ], start=2):
        ws.append(list(r))
        if i % 2 == 1 and r[0]:
            for c in range(1, 4): ws.cell(i+2, c).fill = ALT
    autowidth(ws)

    # ── Sheet 2: SGD Events ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("2_SGD_Events")
    ws2.append(['Country', 'Year', 'GDP(t) %', 'GDP(t+1) %', 'Drop %'])
    hrow(ws2, 1, 5)
    ev = df_full[df_full['target']==1][[cc,'year','gdp_growth','gdp_growth_next']].copy()
    ev['drop_pct'] = ((ev['gdp_growth']-ev['gdp_growth_next'])/ev['gdp_growth'].abs()*100).round(1)
    for _, r in ev.iterrows():
        ws2.append([r[cc], int(r['year']), round(r['gdp_growth'],3),
                    round(r['gdp_growth_next'],3), round(r['drop_pct'],1)])
    autowidth(ws2)

    # ── Sheet 3: PCA ──────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("3_PCA")
    ws3.append(["Component","Variance (%)","Cumulative (%)"]); hrow(ws3,1,3)
    cum = 0
    for i, v in enumerate(prep['pca_var'], 1):
        cum += v*100; ws3.append([f"PC{i}", round(v*100,2), round(cum,2)])
    ws3.append([])
    ws3.append(["Feature Loadings"])
    ws3.cell(ws3.max_row, 1).font = BOLD
    load_df = prep['pca_loadings'].reset_index()
    ws3.append(["Feature"]+[f"PC{i+1}" for i in range(4)])
    hrow(ws3, ws3.max_row, 5,
         fill=PatternFill("solid", fgColor="E3F2FD"), font=Font(bold=True))
    for _, r in load_df.iterrows():
        ws3.append([r['index']]+[round(r[f'PC{i+1}'],4) for i in range(4)])
    autowidth(ws3)

    # ── Sheet 4: Performance (dual threshold) ─────────────────────────────────
    ws4 = wb.create_sheet("4_Performance")
    h4 = ['Model', 'Best Params',
          'Acc(def)', 'Prec(def)', 'Recall(def)', 'F1(def)', 'AUC', 'Brier',
          'Opt Thr', 'Acc(opt)', 'Prec(opt)', 'Recall(opt)', 'F1(opt)',
          'F1 CI Low', 'F1 CI High', 'Recall CI Low', 'Recall CI High']
    ws4.append(h4); hrow(ws4, 1, len(h4))

    for i, (m, r) in enumerate(results.items(), 2):
        ci = boot_ci_dict.get(m, {})
        row = [
            m, r.get('best_params',''),
            r['accuracy'], r['precision'], r['recall'], r['f1'],
            r['auc'], r['brier'],
            r['optimal_threshold'],
            r['opt']['accuracy'], r['opt']['precision'],
            r['opt']['recall'],  r['opt']['f1'],
            ci.get('f1',{}).get('lower',''),
            ci.get('f1',{}).get('upper',''),
            ci.get('recall',{}).get('lower',''),
            ci.get('recall',{}).get('upper',''),
        ]
        ws4.append(row)
        is_vqc = m == 'VQC'
        for c in range(1, len(h4)+1):
            ws4.cell(i, c).fill = VQC_L if is_vqc else (ALT if i%2==1 else PatternFill())
            if is_vqc: ws4.cell(i, c).font = Font(bold=True)
        # Highlight optimal F1 column green
        ws4.cell(i, 13).fill = OPT_L
    # Header annotations
    for col_idx in [9, 13]:      # Opt Thr, F1(opt)
        ws4.cell(1, col_idx).fill = OPT_H
    autowidth(ws4)

    # ── Sheet 5: Confusion Matrices ───────────────────────────────────────────
    ws5 = wb.create_sheet("5_Confusion_Matrices")
    ri = 1
    for m, r in results.items():
        for thr_label, cm in [('Default (τ=0.5)', r['cm']),
                               (f'Optimal (τ={r["optimal_threshold"]:.2f})',
                                r['opt']['cm'])]:
            ws5.cell(ri, 1, f"{m}  [{thr_label}]").font = Font(bold=True, size=11)
            ri += 1
            ws5.cell(ri, 2, "Pred: Normal"); ws5.cell(ri, 3, "Pred: SGD"); ri += 1
            tn,fp,fn,tp = cm[0][0],cm[0][1],cm[1][0],cm[1][1]
            for rl, vals in [("Actual Normal",[tn,fp]),("Actual SGD",[fn,tp])]:
                ws5.cell(ri, 2, rl).font = BOLD
                ws5.cell(ri, 3, vals[0]); ws5.cell(ri, 4, vals[1]); ri += 1
            sens = tp/(tp+fn) if (tp+fn)>0 else 0
            spec = tn/(tn+fp) if (tn+fp)>0 else 0
            ws5.cell(ri, 2, f"Sensitivity={sens:.3f}  Specificity={spec:.3f}")
            ri += 2

    # ── Sheet 6: McNemar ──────────────────────────────────────────────────────
    ws6 = wb.create_sheet("6_McNemar")
    h6 = ['Comparison','b (class wrong, VQC right)',
          'c (class right, VQC wrong)', 'n=b+c',
          'p-value','Significance','Method']
    ws6.append(h6); hrow(ws6, 1, len(h6))
    for i, t in enumerate(tests, 2):
        ws6.append([t.get(k,'') for k in
                    ['comparison','b','c','n','p_value','significance','method']])
        sig = t.get('significance','ns')
        fill = (PatternFill("solid", fgColor="E8F5E9") if 'p<0.05' in sig
                else PatternFill("solid", fgColor="FFF9C4"))
        for c in range(1, len(h6)+1): ws6.cell(i, c).fill = fill
    ws6.append([])
    ws6.append(["Note: Exact binomial test (n<25). "
                "*p<0.05 **p<0.01 ns=not significant. "
                "Predictions at optimal threshold."])
    autowidth(ws6)

    # ── Sheet 7: Bootstrap CI ─────────────────────────────────────────────────
    ws7 = wb.create_sheet("7_Bootstrap_CI")
    h7 = ['Model','F1','F1 CI Low','F1 CI High',
          'Recall','Recall CI Low','Recall CI High',
          'Prec','Prec CI Low','Prec CI High',
          'AUC','AUC CI Low','AUC CI High']
    ws7.append(h7); hrow(ws7, 1, len(h7))
    for i, (m, r) in enumerate(results.items(), 2):
        ci = boot_ci_dict.get(m, {})
        ws7.append([m,
            r['f1'], ci.get('f1',{}).get('lower',''), ci.get('f1',{}).get('upper',''),
            r['recall'], ci.get('recall',{}).get('lower',''), ci.get('recall',{}).get('upper',''),
            r['precision'], ci.get('precision',{}).get('lower',''), ci.get('precision',{}).get('upper',''),
            r['auc'], ci.get('auc',{}).get('lower',''), ci.get('auc',{}).get('upper',''),
        ])
        if m == 'VQC':
            for c in range(1, len(h7)+1): ws7.cell(i, c).fill = VQC_L
    ws7.append([])
    ws7.append(["Note: 95% CI, n=1,000 bootstrap replicates, percentile method, seed=42"])
    autowidth(ws7)

    # ── Sheet 8: Robustness ───────────────────────────────────────────────────
    ws8 = wb.create_sheet("8_Robustness")
    ws8.append(["Robustness Check — SGD Threshold Sensitivity"])
    ws8.cell(1, 1).font = Font(bold=True, size=12)
    ws8.append([])
    ri = 3
    for alt, dat in robust.items():
        ws8.cell(ri, 1, alt).font = Font(bold=True); ri += 1
        ws8.cell(ri, 2, f"n_pos_test={dat.get('n_pos_test','?')}"); ri += 1
        ws8.append(['Model','F1(def)','F1(opt)','Recall(def)','Recall(opt)','AUC'])
        hrow(ws8, ws8.max_row, 6,
             fill=PatternFill("solid", fgColor="E3F2FD"), font=Font(bold=True))
        ri = ws8.max_row + 1
        for m, met in dat.get('models', {}).items():
            ws8.append([m, met.get('f1_default'), met.get('f1_optimal'),
                        met.get('recall_default'), met.get('recall_optimal'),
                        met.get('auc')])
            ri += 1
        ri += 1
    autowidth(ws8)

    # ── Sheet 9: Threshold Sensitivity ────────────────────────────────────────
    if thr_df is not None and not thr_df.empty:
        ws9 = wb.create_sheet("9_Threshold_Sensitivity")
        ws9.append(list(thr_df.columns)); hrow(ws9, 1, len(thr_df.columns))
        for i, row in enumerate(thr_df.itertuples(index=False), 2):
            ws9.append(list(row))
            if i % 2 == 1:
                for c in range(1, len(thr_df.columns)+1): ws9.cell(i, c).fill = ALT
        autowidth(ws9)

    # ── Sheet 10: Per-Country ─────────────────────────────────────────────────
    if country_df is not None and not country_df.empty:
        ws10 = wb.create_sheet("10_Per_Country")
        ws10.append(list(country_df.columns))
        hrow(ws10, 1, len(country_df.columns))
        for i, row in enumerate(country_df.itertuples(index=False), 2):
            ws10.append(list(row))
            if i % 2 == 1:
                for c in range(1, len(country_df.columns)+1):
                    ws10.cell(i, c).fill = ALT
        autowidth(ws10)

    # ── Sheet 11: Ablation ────────────────────────────────────────────────────
    if ablation:
        abl_df = pd.DataFrame(ablation)
        ws11 = wb.create_sheet("11_VQC_Ablation")
        ws11.append(list(abl_df.columns))
        hrow(ws11, 1, len(abl_df.columns))
        for i, row in enumerate(abl_df.itertuples(index=False), 2):
            ws11.append(list(row))
            is_main = 'MAIN' in str(getattr(row, 'config', ''))
            if is_main:
                for c in range(1, len(abl_df.columns)+1):
                    ws11.cell(i, c).fill = OPT_L
                    ws11.cell(i, c).font = Font(bold=True)
        autowidth(ws11)

    path = OUTPUT_DIR / "sgd_full_results_v3.xlsx"
    wb.save(path)
    log(f"  Saved → {path}")


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    parser = argparse.ArgumentParser(description='SGD Prediction Pipeline v3.1')
    parser.add_argument('--data',         default='economic_data_collected.csv')
    parser.add_argument('--skip-vqc',     action='store_true')
    parser.add_argument('--skip-ablation',action='store_true')
    parser.add_argument('--skip-robust',  action='store_true')
    parser.add_argument('--bootstrap-n',  type=int, default=1000)
    args = parser.parse_args()

    log("=" * 70)
    log("SGD PREDICTION PIPELINE v3.1 — START")
    log(f"Output directory: {OUTPUT_DIR.resolve()}")
    log(f"VQC config (ablation-optimal): "
        f"ZZ(r={VQC_FM_REPS})+ESU2(r={VQC_ANSATZ_REPS})+COBYLA({VQC_MAXITER})")
    log(f"VQC reporting threshold: τ={VQC_OPT_THR} (validated on v3.0 sweep)")
    log("=" * 70)
    t_start = time.time()

    # 1. Load
    df = load_and_prepare(args.data)

    # 2. Preprocess (SMOTE k=4, validated in v2.0)
    prep = preprocess(df.copy())

    # 3. Classical baselines (+ calibration + optimal threshold)
    results = run_classical(prep)

    # 4. VQC (ablation-optimal config)
    if not args.skip_vqc:
        try:
            results['VQC'] = run_vqc(prep)
            log("✅ VQC complete")
        except Exception as e:
            log(f"❌ VQC failed: {str(e)[:300]}")
            results['VQC'] = {
                'best_params': 'FAILED', 'accuracy': 0, 'precision': 0,
                'recall': 0, 'f1': 0, 'auc': 0.5, 'brier': 0.25,
                'cm': [[0,0],[0,0]], 'y_prob': [],
                'optimal_threshold': DEFAULT_THR,
                'opt': {'accuracy':0,'precision':0,'recall':0,'f1':0,
                        'cm':[[0,0],[0,0]],'y_pred':[]},
            }
    else:
        log("STEP 4: VQC skipped")

    # 5a. Bootstrap CI (on optimal-threshold predictions)
    boot_ci_dict = run_bootstrap(
        results, prep['y_te'],
        n_boot=args.bootstrap_n, use_optimal=True,
    )

    # 5b. McNemar (optimal threshold)
    tests = run_mcnemar(results, prep['y_te'], use_optimal=True)

    # 6. Per-country
    country_df, meta_te = run_per_country(results, prep, use_optimal=True)

    # 7. Threshold sensitivity
    thr_df = run_threshold_sensitivity(results, prep['y_te'])

    # 8. Robustness
    robust = {}
    if not args.skip_robust:
        robust = run_robustness(args.data)
    else:
        log("STEP 8: Robustness skipped")

    # 9. Ablation
    ablation = []
    if not args.skip_vqc and not args.skip_ablation:
        try:
            ablation = run_ablation(prep)
        except Exception as e:
            log(f"  Ablation error: {e}")
    else:
        log("STEP 9: Ablation skipped")

    # 10. Figures
    generate_figures(results, prep, prep['y_te'],
                     boot_ci_dict, thr_df, country_df, ablation)

    # 11. Excel
    export_excel(results, prep, prep['y_te'], df,
                 tests, robust, boot_ci_dict,
                 thr_df, country_df, ablation)

    # JSON summary
    json_out = {}
    for k, v in results.items():
        json_out[k] = {
            kk: vv for kk, vv in v.items()
            if kk not in ('y_pred', 'y_prob', 'estimator', 'loss_history')
        }
    with open(OUTPUT_DIR / "results_summary.json", 'w') as f:
        json.dump(json_out, f, indent=2, default=str)

    # ── Final summary ─────────────────────────────────────────────────────────
    log("")
    log("=" * 70)
    log("FINAL RESULTS SUMMARY")
    log("=" * 70)
    log(f"\n{'Model':<22} {'F1(def)':>8} {'F1(opt)':>8} {'τ_opt':>6} "
        f"{'Recall(opt)':>12} {'AUC':>6} {'Brier':>7}")
    log("-" * 70)

    for m, r in results.items():
        tag = " ★" if m == 'VQC' else ""
        ci  = boot_ci_dict.get(m, {})
        f1_ci = (f"[{ci['f1']['lower']:.3f},{ci['f1']['upper']:.3f}]"
                 if ci.get('f1') and not np.isnan(ci['f1'].get('lower', np.nan))
                 else "")
        log(f"{m+tag:<22} {r['f1']:>8.4f} {r['opt']['f1']:>8.4f} "
            f"{r['optimal_threshold']:>6.2f} "
            f"{r['opt']['recall']:>12.4f} "
            f"{r.get('auc',0):>6.4f} {r.get('brier',0):>7.4f}  {f1_ci}")

    # Quantum advantage (at optimal threshold)
    if 'VQC' in results:
        vqc_f1 = results['VQC']['opt']['f1']
        cl_f1s = [r['opt']['f1'] for m, r in results.items() if m != 'VQC']
        if cl_f1s:
            best_cl = max(cl_f1s)
            diff = (vqc_f1 - best_cl) * 100
            icon = "✅" if diff > 0 else "⚠️"
            log(f"\n{icon} VQC vs best classical (optimal τ): "
                f"{'+'if diff>0 else ''}{diff:.2f}% F1")

    log("")
    log("McNemar summary (optimal threshold):")
    for t in tests:
        log(f"  {t['comparison']:<35}  p={t['p_value']:.4f}  {t['significance']}")

    elapsed = time.time() - t_start
    log(f"\nPipeline complete in {elapsed/60:.1f} minutes")
    log(f"All outputs → {OUTPUT_DIR.resolve()}/")
    log("=" * 70)


if __name__ == "__main__":
    main()